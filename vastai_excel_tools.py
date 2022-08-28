'''
Author: liu kang
Date: 2022-08-28 19:14:36
LastEditors: fade_away_shot
LastEditTime: 2022-08-28 23:33:41
FilePath: \pythonTool\vastai_excel_tools.py
Description: 

Copyright (c) 2022 by fade_away_shot 654042016@qq.com, All Rights Reserved. 
'''
# https://www.h3blog.com/article/229/

from openpyxl import load_workbook, Workbook
import datetime

def get_data_file_name(_filename:str)->str:
    _dayTime =datetime.datetime.now().strftime('%Y-%m-%d')
    _hourTime = datetime.datetime.now().strftime('%H_%M_%S')
    return _dayTime+'-'+_hourTime+_filename 


class Excel_manager():
    def __init__(self, _file_list:list, _sheet:str) -> None:
        self._file_number = len(_file_list)
        self._file_list = _file_list
        self._wb = None
        self._list_data = []
        self._file_title = []
        self._sheet_name = _sheet
        self._check_result =[]
    def load_excel(self):
        for i, _file in enumerate(self._file_list):
            _wb = load_workbook(_file)
            _ws = _wb[self._sheet_name]
            for k, _value in enumerate(_ws.values):
                if 0==k :
                    self._file_title = [v for v in _value]
                else:
                    self._list_data.append([v for v in _value])
            _wb.close()
    def check_history_result(self):
            pass
    
    def check_self_result(self):
            pass
        
    def get_new_excel(self,_dst_name:str, _dst_path=None):
        _wb = Workbook()
        _wb.create_sheet(self._sheet_name)
        title = self._file_title
        _ws1 = _wb[self._sheet_name]
        _ws1.append(title)
        del _wb['Sheet']
        _ws1.append(self._check_result)
        # _out_put = None
        # if not _dst_path:
        _out_put = _dst_name + _dst_path if  _dst_path !=None  else _dst_name
        _wb.save(_out_put)
    
        
if __name__ == '__main__':
    file = ['D:/data/2022-05-25-10-30-22-endurance-cls.xlsx']
    a = Excel_manager(file, 'accuracy')
    a.load_excel()
    a.get_new_excel('bbb.xlsx')